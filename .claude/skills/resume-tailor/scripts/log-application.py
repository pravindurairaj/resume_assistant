#!/usr/bin/env python3
"""
log-application.py - Append a job application record to {User}/History/resumes_created.xlsx.

Creates the file and directory if they don't exist.
Appends one row per invocation — never overwrites existing data.

Usage:
    python log-application.py \\
        --user Pravin \\
        --company "Acme Corp" \\
        --title "Senior Data Analyst" \\
        --location "Dublin, Ireland" \\
        --link "https://www.linkedin.com/jobs/view/12345" \\
        --resume-file "Pravin/Resumes/AcmeCorp/Pravin_Resume_AcmeCorp.docx" \\
        --fit 85

Requirements:
    pip install openpyxl
    (or: pip install -r .github/requirements.txt)
"""

import argparse
import sys
from datetime import date
from pathlib import Path


# ── Dependency check ──────────────────────────────────────────────────────────

def check_dependencies():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERROR: Missing dependency: openpyxl")
        print("Run:  pip install -r .github/requirements.txt")
        sys.exit(1)


# ── Excel constants ────────────────────────────────────────────────────────────

HEADERS    = ["Date Applied", "Company", "Job Title", "Location", "Job Link", "Resume File", "Fit %"]
COL_WIDTHS = [14,             28,        38,          20,         18,         55,            8]


# ── Workbook creation ─────────────────────────────────────────────────────────

def create_workbook(xlsx_path: Path):
    """Create a new applied_jobs.xlsx with a styled navy header row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Applied Jobs"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin     = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    ws.append(HEADERS)

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin
        ws.column_dimensions[c.column_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    wb.close()


# ── Row append ────────────────────────────────────────────────────────────────

def append_row(xlsx_path: Path, row_data: dict):
    """Append a single data row to an existing applied_jobs.xlsx."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = load_workbook(xlsx_path)
    ws = wb.active

    next_row = ws.max_row + 1
    thin = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    # Alternating row fill for readability
    if next_row % 2 == 0:
        row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    else:
        row_fill = PatternFill(fill_type=None)

    link_url = row_data.get("link", "")

    column_keys = ["date", "company", "title", "location", "link", "resume_file", "fit"]
    for col_idx, key in enumerate(column_keys, 1):
        val = row_data.get(key, "")
        c   = ws.cell(row=next_row, column=col_idx, value=val)
        c.border = thin

        if row_fill.fill_type:
            c.fill = row_fill

        # Job Link column → clickable hyperlink
        if col_idx == 5 and link_url:
            c.value     = "Open ↗"
            c.hyperlink = link_url
            c.font      = Font(color="1F4E79", underline="single")
            c.alignment = Alignment(horizontal="center")

        # Fit % column → centred, bold
        elif col_idx == 7:
            c.alignment = Alignment(horizontal="center")
            c.font      = Font(bold=True)

    wb.save(xlsx_path)
    wb.close()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Append a job application record to {User}/History/resumes_created.xlsx."
    )
    parser.add_argument("--user",        "-u", required=True, help="User name (e.g. Pravin)")
    parser.add_argument("--company",     "-c", required=True, help="Company name")
    parser.add_argument("--title",       "-t", required=True, help="Job title")
    parser.add_argument("--location",    "-l", required=True, help="Job location")
    parser.add_argument("--link",              required=True, help="LinkedIn job URL")
    parser.add_argument("--resume-file", "-r", required=True,
                        help="Relative path to the generated .docx resume")
    parser.add_argument("--fit",         "-f", type=int, required=True,
                        help="Fit percentage as an integer (e.g. 85)")
    args = parser.parse_args()

    check_dependencies()

    xlsx_path = Path(args.user) / "History" / "resumes_created.xlsx"

    if not xlsx_path.exists():
        print(f"Creating {xlsx_path} ...")
        create_workbook(xlsx_path)

    row_data = {
        "date":        date.today().isoformat(),
        "company":     args.company,
        "title":       args.title,
        "location":    args.location,
        "link":        args.link,
        "resume_file": args.resume_file,
        "fit":         f"{args.fit}%",
    }

    append_row(xlsx_path, row_data)
    print(f"Logged: {args.company} — {args.title} ({args.fit}%) → {xlsx_path}")


if __name__ == "__main__":
    main()

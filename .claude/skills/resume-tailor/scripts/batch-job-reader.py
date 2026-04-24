#!/usr/bin/env python3
"""
batch-job-reader.py - Read latest JobSearch Excel, filter by Fit %, produce JSON manifest.

Reads {User}/JobSearch/*.xlsx (or a specific file), filters jobs by Fit %, checks for
existing resumes, loads any pre-saved JD text files from {User}/JobSearch/jds/, and
outputs a JSON manifest for the resume-tailor agent batch loop.

JD text is auto-scraped via the LinkedIn guest API when available. Pre-saved JD text
files in {User}/JobSearch/jds/{folder_name}.txt take priority. Any remaining jobs
without JDs will be scraped automatically before manifest generation.

Usage:
    # Auto — latest Excel for Pravin, Fit >= 50%
    python batch-job-reader.py --user Pravin

    # Stricter fit threshold
    python batch-job-reader.py --user Pravin --min-fit 75

    # Resume an interrupted batch (skip already-completed jobs)
    python batch-job-reader.py --user Pravin --manifest Pravin/JobSearch/batch_manifest_20260411.json

    # Specific Excel file, cap at 10 jobs
    python batch-job-reader.py --user Pravin --excel Pravin/JobSearch/specific.xlsx --max-jobs 10

Requirements:
    pip install openpyxl
    (or: pip install -r .github/requirements.txt)
"""

import argparse
import importlib.util
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path


# ── Dependency check ──────────────────────────────────────────────────────────

def check_dependencies():
    missing = []
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print(f"ERROR: Missing dependencies: {', '.join(missing)}")
        print("Run:  pip install -r .github/requirements.txt")
        sys.exit(1)


# ── Repo root ────────────────────────────────────────────────────────────────

def find_repo_root():
    """Walk up from script location to find the repo root (contains .github/)."""
    path = Path(__file__).resolve().parent
    while path != path.parent:
        if (path / ".github").is_dir():
            return path
        path = path.parent
    return None


# ── Career profile ────────────────────────────────────────────────────────────

def load_career_profile(user_name, repo_root):
    """
    Load full_name, location, and target_roles from career-profile-{user}.instructions.md.
    Returns dict; returns empty dict if profile not found.
    """
    instructions_dir = repo_root / ".github" / "instructions"
    if not instructions_dir.exists():
        return {}

    for f in instructions_dir.glob("career-profile-*.instructions.md"):
        if user_name.lower() in f.stem.lower():
            content = f.read_text(encoding="utf-8")
            profile = {"full_name": "", "location": "", "target_roles": []}
            current_section = ""

            for line in content.splitlines():
                stripped = line.strip()

                if stripped.startswith("## "):
                    current_section = stripped[3:].strip().lower()
                    continue

                if current_section == "contact information" and "|" in stripped:
                    parts = [p.strip() for p in stripped.split("|")]
                    for i, part in enumerate(parts):
                        if "Location" in part and i + 1 < len(parts):
                            loc = parts[i + 1].strip().strip("*")
                            if loc and not loc.startswith("{"):
                                profile["location"] = loc
                        if "Full Name" in part and i + 1 < len(parts):
                            name = parts[i + 1].strip().strip("*")
                            if name and not name.startswith("{"):
                                profile["full_name"] = name

                if current_section == "target roles" and stripped.startswith("- "):
                    role = stripped[2:].strip()
                    if role and not role.startswith("{"):
                        profile["target_roles"].append(role)

            return profile

    return {}


# ── Name sanitization ─────────────────────────────────────────────────────────

def sanitize_company(name):
    """
    Normalize a company name to a folder-safe alphanumeric string.
    Strips common corporate suffixes (Ltd, Inc, Corp, etc.).
    'Acme Corp Ltd.' → 'Acme'   |   'A&B Solutions' → 'ABSolutions'
    """
    name = re.sub(
        r'\b(Ltd|Limited|Inc|Incorporated|Corp|Corporation|GmbH|PLC|LLC|Co)\b\.?',
        '', name, flags=re.IGNORECASE
    )
    sanitized = re.sub(r'[^a-zA-Z0-9]', '', name.strip())
    return sanitized or "Unknown"


def shorten_title(title):
    """
    Shorten a job title to a compact CamelCase string for use in folder/file names.
    'Senior Data Analyst' → 'DataAnalyst'   |   'Lead Platform Engineer' → 'PlatformEngineer'
    """
    title = re.sub(
        r'\b(Senior|Sr\.?|Junior|Jr\.?|Lead|Staff|Principal|Associate|Snr\.?|Head of)\b',
        '', title, flags=re.IGNORECASE
    )
    words = re.findall(r'[a-zA-Z0-9]+', title)
    return ''.join(w.capitalize() for w in words) or "Role"


# ── Excel parsing ─────────────────────────────────────────────────────────────

def find_latest_excel(jobsearch_dir):
    """Return the most recently modified .xlsx in jobsearch_dir, or None."""
    files = [f for f in jobsearch_dir.glob("*.xlsx") if not f.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda f: f.stat().st_mtime)


def read_excel_jobs(xlsx_path, min_fit):
    """
    Parse a LinkedIn job scraper Excel file and return filtered job list.

    Excel structure produced by scrape-linkedin-jobs.py:
      Row 1  : Title metadata
      Row 2  : Timestamp metadata
      Row 3  : Spacer
      Row 4  : Column headers  →  #, Company, Job Title, Location, Job Link, [Fit %], [Best Matching Role]
      Row 5+ : Data rows

    Job Link (column E) is stored as a hyperlink — the displayed value is "Open Job ↗"
    and the actual URL is in cell.hyperlink.target.

    Fit % column (F) is optional — only present when scraper had target_roles configured.
    If absent, all jobs are included regardless of min_fit.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    HEADER_ROW = 4
    DATA_START  = HEADER_ROW + 1

    # Read headers to detect optional columns
    headers = [
        (ws.cell(row=HEADER_ROW, column=c).value or "")
        for c in range(1, ws.max_column + 1)
    ]

    # Fit % is column F (index 5, 0-based) — detect by header text
    has_fit_col = (
        len(headers) >= 6
        and "fit" in str(headers[5]).lower()
    )

    jobs = []

    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        # Column B → Company (index 1)
        company_cell = row[1] if len(row) > 1 else None
        if not company_cell or not company_cell.value:
            continue

        company  = str(company_cell.value or "").strip()
        title    = str(row[2].value or "").strip() if len(row) > 2 else ""
        location = str(row[3].value or "").strip() if len(row) > 3 else ""

        # Column E → Job Link (index 4): read hyperlink target, not display value
        link = ""
        if len(row) > 4:
            link_cell = row[4]
            if link_cell.hyperlink is not None:
                # hyperlink.target is the real URL; display value is "Open Job ↗"
                raw = str(getattr(link_cell.hyperlink, "target", "") or "")
                link = raw.split("?")[0].strip()
            elif link_cell.value and str(link_cell.value).startswith("http"):
                # Fallback: plain text URL stored in cell value
                link = str(link_cell.value).split("?")[0].strip()

        # Column F → Fit % (index 5, optional)
        fit_pct = 0
        if has_fit_col and len(row) > 5 and row[5].value is not None:
            try:
                fit_pct = int(str(row[5].value).replace("%", "").strip())
            except ValueError:
                fit_pct = 0

        # Column G → Best Matching Role (index 6, optional)
        matching_role = ""
        if len(row) > 6 and row[6].value:
            matching_role = str(row[6].value).strip()

        if not company or not title:
            continue

        # Apply fit filter only when the column exists
        if has_fit_col and fit_pct < min_fit:
            continue

        jobs.append({
            "company":       company,
            "title":         title,
            "location":      location,
            "link":          link,
            "fit_pct":       fit_pct,
            "matching_role": matching_role,
        })

    wb.close()
    return jobs


# ── Output path resolution ────────────────────────────────────────────────────

def resolve_output_names(jobs, user, repo_root):
    """
    Compute folder_name, output_folder, output_md, output_docx, context_file for each job.

    Rules:
    - Single job for company   → folder = {CompanySanitized}
    - Multiple jobs per company → folder = {CompanySanitized}_{TitleShort}
    - Duplicate company+title  → folder = {CompanySanitized}_{TitleShort}_{N}
    - Sanitized-name collision (different companies → same string) → {CompanySanitized}_{hash_suffix}
    """
    # Build sanitized forms
    sanitized_list = [sanitize_company(j["company"]) for j in jobs]
    title_list     = [shorten_title(j["title"])      for j in jobs]

    # Count occurrences of (sanitized_company, title_short) pairs
    count_by_company = {}
    for s in sanitized_list:
        count_by_company[s] = count_by_company.get(s, 0) + 1

    count_by_key = {}
    for s, t in zip(sanitized_list, title_list):
        key = (s, t)
        count_by_key[key] = count_by_key.get(key, 0) + 1

    key_seen = {}
    results  = []

    for job, cs, ts in zip(jobs, sanitized_list, title_list):
        key = (cs, ts)

        if count_by_key[key] > 1:
            # Multiple identical company+title combos → add numeric suffix
            key_seen[key] = key_seen.get(key, 0) + 1
            folder_name = f"{cs}_{ts}_{key_seen[key]}"
        elif count_by_company[cs] > 1:
            # Same company, different roles → include title
            folder_name = f"{cs}_{ts}"
        else:
            folder_name = cs

        resumes_dir  = Path(user) / "Resumes" / folder_name
        output_md    = repo_root / ".github" / "Users" / user / f"{user}_Resume_{folder_name}.md"
        output_docx  = resumes_dir / f"{user}_Resume_{folder_name}.docx"
        context_file = repo_root / ".github" / "Users" / user / "companies" / f"{folder_name}.md"

        results.append({
            **job,
            "company_sanitized": cs,
            "title_short":       ts,
            "folder_name":       folder_name,
            "output_folder":     str(resumes_dir),
            "output_md":         str(output_md),
            "output_docx":       str(output_docx),
            "context_file":      str(context_file),
        })

    return results


# ── Skip checks ───────────────────────────────────────────────────────────────

def check_existing_resumes(jobs, user):
    """
    Mark a job as skipped if its output Resumes subfolder already contains a .docx.
    Returns (to_process, skipped_list).
    """
    to_process = []
    skipped    = []

    for job in jobs:
        folder = Path(job["output_folder"])
        if folder.exists() and any(folder.glob("*.docx")):
            skipped.append({
                **{k: job[k] for k in ("company", "title", "link", "fit_pct", "folder_name")},
                "reason": f"resume exists at {job['output_folder']}",
            })
        else:
            to_process.append(job)

    return to_process, skipped


def load_manifest_completed(manifest_path):
    """
    Read a previous manifest JSON and return a set of folder_names with status=completed.
    Used for resuming an interrupted batch.
    """
    try:
        data = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
        return {
            j["folder_name"]
            for j in data.get("jobs", [])
            if j.get("status") == "completed"
        }
    except Exception as e:
        print(f"WARNING: Could not read manifest '{manifest_path}': {e}")
        return set()


# ── JD pre-loading ────────────────────────────────────────────────────────────

def load_jd_files(jobs, user):
    """
    Check {User}/JobSearch/jds/ for pre-saved JD text files and load them.

    File matching order (first match wins):
      1. {job_id}.txt  — saved by scrape-linkedin-jobs.py during scrape phase
      2. {folder_name}.txt
      3. {company_sanitized}_{title_short}.txt
      4. {company_sanitized}.txt

    Sets jd_text and jd_source on each job dict.
    """
    import re
    jds_dir = Path(user) / "JobSearch" / "jds"

    for job in jobs:
        jd_text   = ""
        jd_source = "pending_user_input"

        if jds_dir.exists():
            # job_id from link (scraper saves as {job_id}.txt)
            job_id_match = re.search(r'(\d{8,})', job.get("link") or "")
            job_id_file  = jds_dir / f"{job_id_match.group(1)}.txt" if job_id_match else None

            candidates = [c for c in [
                job_id_file,
                jds_dir / f"{job['folder_name']}.txt",
                jds_dir / f"{job['company_sanitized']}_{job['title_short']}.txt",
                jds_dir / f"{job['company_sanitized']}.txt",
            ] if c is not None]

            for candidate in candidates:
                if candidate.exists():
                    try:
                        jd_text   = candidate.read_text(encoding="utf-8").strip()
                        jd_source = f"file:{candidate.name}"
                        break
                    except Exception:
                        pass

        job["jd_text"]   = jd_text
        job["jd_source"] = jd_source

    return jobs


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Batch job reader — produce resume-tailor manifest from latest JobSearch Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--user",     "-u", default="Pravin",
                        help="User name (matches career-profile-{user}.instructions.md). Default: Pravin")
    parser.add_argument("--min-fit",  "-f", type=int, default=50,
                        help="Minimum Fit %% to include (default: 50). Ignored if Excel has no Fit %% column.")
    parser.add_argument("--excel",    "-e", default=None,
                        help="Specific .xlsx file to read. Default: latest in {User}/JobSearch/")
    parser.add_argument("--output",   "-o", default=None,
                        help="Manifest output path. Default: {User}/JobSearch/batch_manifest_{ts}.json")
    parser.add_argument("--max-jobs", "-m", type=int, default=None,
                        help="Cap number of jobs to process (after filtering and skipping)")
    parser.add_argument("--manifest",       default=None,
                        help="Path to a previous manifest JSON — skips jobs with status=completed")
    args = parser.parse_args()

    check_dependencies()

    repo_root = find_repo_root()
    if not repo_root:
        print("ERROR: Could not find repo root (looking for a parent directory with .github/)")
        sys.exit(1)

    user          = args.user
    jobsearch_dir = Path(user) / "JobSearch"

    if not jobsearch_dir.exists():
        print(f"ERROR: {jobsearch_dir} does not exist. Run the job scraper first.")
        sys.exit(1)

    # ── Career profile ─────────────────────────────────────────────────────────
    profile        = load_career_profile(user, repo_root)
    user_full_name = profile.get("full_name") or user

    # ── Find / validate Excel ──────────────────────────────────────────────────
    xlsx_path = Path(args.excel) if args.excel else find_latest_excel(jobsearch_dir)
    if not xlsx_path or not xlsx_path.exists():
        print(f"ERROR: No .xlsx file found in {jobsearch_dir}. Run the job scraper first.")
        sys.exit(1)

    print(f"\n{'─' * 60}")
    print(f"  batch-job-reader.py")
    print(f"{'─' * 60}")
    print(f"  User         : {user} ({user_full_name})")
    print(f"  Source Excel : {xlsx_path.name}")
    print(f"  Min Fit %%    : {args.min_fit}%%")

    # ── Parse Excel ────────────────────────────────────────────────────────────
    try:
        all_jobs = read_excel_jobs(xlsx_path, args.min_fit)
    except Exception as e:
        print(f"\nERROR: Failed to parse Excel file: {e}")
        sys.exit(1)

    total_after_filter = len(all_jobs)
    print(f"  After filter : {total_after_filter} job(s) with Fit ≥ {args.min_fit}%%")

    if not all_jobs:
        print("\n  No jobs meet the fit threshold. Try a lower --min-fit value.")
        sys.exit(0)

    # ── Resolve output paths ───────────────────────────────────────────────────
    all_jobs = resolve_output_names(all_jobs, user, repo_root)

    # ── Skip: resume from previous manifest ───────────────────────────────────
    completed_folders = set()
    skipped_prev_run  = []
    if args.manifest:
        completed_folders = load_manifest_completed(args.manifest)
        if completed_folders:
            print(f"  Resuming     : {len(completed_folders)} job(s) already completed in previous run")

    # ── Skip: existing .docx in Resumes/ ──────────────────────────────────────
    to_process, skipped_existing = check_existing_resumes(all_jobs, user)

    if completed_folders:
        remaining = []
        for job in to_process:
            if job["folder_name"] in completed_folders:
                skipped_prev_run.append({
                    **{k: job[k] for k in ("company", "title", "link", "fit_pct", "folder_name")},
                    "reason": "completed in previous manifest run",
                })
            else:
                remaining.append(job)
        to_process = remaining

    # ── Cap --max-jobs ─────────────────────────────────────────────────────────
    if args.max_jobs and len(to_process) > args.max_jobs:
        print(f"  Capping to {args.max_jobs} job(s) (--max-jobs)")
        to_process = to_process[:args.max_jobs]

    # ── Load pre-saved JD files ────────────────────────────────────────────────
    to_process = load_jd_files(to_process, user)
    # ── Scrape JDs from LinkedIn guest API for remaining jobs ──────────────────
    scraper_path = repo_root / ".github" / "skills" / "job-scraper" / "scripts" / "scrape-linkedin-jobs.py"
    scraper_module = None
    if scraper_path.exists():
        spec = importlib.util.spec_from_file_location("scraper", scraper_path)
        scraper_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(scraper_module)

    pending_jd = [j for j in to_process if j["jd_source"] == "pending_user_input"]
    if pending_jd and scraper_module:
        import re as _re
        jds_dir = Path(user) / "JobSearch" / "jds"
        jds_dir.mkdir(parents=True, exist_ok=True)
        print(f"\n  Scraping JDs from LinkedIn for {len(pending_jd)} job(s)...")
        for idx, job in enumerate(pending_jd, 1):
            if job["link"]:
                label = f"{job['company'][:28]} — {job['title'][:32]}"
                print(f"    [{idx}/{len(pending_jd)}] {label}...", end=" ", flush=True)
                jd_text, ok = scraper_module.scrape_job_description(job["link"])
                if ok and jd_text:
                    job["jd_text"] = jd_text
                    job["jd_source"] = "scraped_linkedin"
                    # Save for reuse by future runs
                    jd_file = jds_dir / f"{job['folder_name']}.txt"
                    try:
                        jd_file.write_text(jd_text, encoding="utf-8")
                    except Exception:
                        pass
                    print("\u2713")
                else:
                    print("\u2717 (blocked/empty)")
                if idx < len(pending_jd):
                    time.sleep(5)
    jd_loaded  = sum(1 for j in to_process if j["jd_source"] != "pending_user_input")
    jd_pending = len(to_process) - jd_loaded

    # ── Build manifest ─────────────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    manifest_jobs = []
    for idx, job in enumerate(to_process, 1):
        manifest_jobs.append({
            "index":             idx,
            "company":           job["company"],
            "company_sanitized": job["company_sanitized"],
            "title":             job["title"],
            "title_short":       job["title_short"],
            "folder_name":       job["folder_name"],
            "location":          job["location"],
            "link":              job["link"],
            "fit_pct":           job["fit_pct"],
            "matching_role":     job["matching_role"],
            "jd_text":           job["jd_text"],
            "jd_source":         job["jd_source"],
            "output_folder":     job["output_folder"],
            "output_md":         job["output_md"],
            "output_docx":       job["output_docx"],
            "context_file":      job["context_file"],
            "status":            "pending",
        })

    all_skipped = skipped_existing + skipped_prev_run

    manifest = {
        "user":              user,
        "user_full_name":    user_full_name,
        "source_excel":      str(xlsx_path),
        "master_resume":     str(repo_root / ".github" / "Users" / user / f"{user}_Resume.md"),
        "generated":         datetime.now().isoformat(timespec="seconds"),
        "filter_min_fit":    args.min_fit,
        "jds_folder_hint":   str(Path(user) / "JobSearch" / "jds"),
        "batch_size_hint":   5,
        "summary": {
            "total_after_fit_filter":      total_after_filter,
            "skipped_existing_resume":     len(skipped_existing),
            "skipped_previous_run":        len(skipped_prev_run),
            "to_process":                  len(to_process),
            "jd_preloaded_from_file":      jd_loaded,
            "jd_pending_user_input":       jd_pending,
        },
        "jobs":    manifest_jobs,
        "skipped": all_skipped,
    }

    # ── Write manifest ─────────────────────────────────────────────────────────
    if args.output:
        manifest_path = Path(args.output)
    else:
        manifest_path = jobsearch_dir / f"batch_manifest_{ts}.json"

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    # ── Print summary ──────────────────────────────────────────────────────────
    print(f"\n{'─' * 60}")
    print(f"  SUMMARY")
    print(f"{'─' * 60}")
    print(f"  Jobs in Excel (Fit ≥{args.min_fit}%%) : {total_after_filter}")
    print(f"  Skipped (resume exists)     : {len(skipped_existing)}")
    print(f"  Skipped (prev run complete) : {len(skipped_prev_run)}")
    print(f"  To process                  : {len(to_process)}")
    print(f"  JD pre-loaded from file     : {jd_loaded}")
    print(f"  JD pending (user paste)     : {jd_pending}")
    print(f"{'─' * 60}")
    print(f"  Manifest: {manifest_path}")
    print(f"{'─' * 60}")

    if jd_pending > 0:
        jds_dir = Path(user) / "JobSearch" / "jds"
        print(f"\n  NOTE: {jd_pending} job(s) need JD text.")
        print(f"  Option A — Save JD files before running:")
        print(f"    Folder: {jds_dir}")
        print(f"    Files:  {{folder_name}}.txt  (see 'folder_name' in manifest)")
        print(f"  Option B — Paste JDs interactively in the agent chat.\n")

    if manifest_jobs:
        print(f"\n  Jobs to process:")
        print(f"  {'#':>3}  {'Company':<28}  {'Title':<32}  {'Fit':>4}  JD")
        print(f"  {'─'*3}  {'─'*28}  {'─'*32}  {'─'*4}  {'─'*12}")
        for j in manifest_jobs:
            jd_flag = "✓ ready" if j["jd_source"] != "pending_user_input" else "⏳ needed"
            company_col = j["company"][:28]
            title_col   = j["title"][:32]
            print(f"  {j['index']:>3}  {company_col:<28}  {title_col:<32}  {j['fit_pct']:>3}%  {jd_flag}")

    print(f"\n  Pass to agent:\n    read the file at:  {manifest_path}\n")


if __name__ == "__main__":
    main()

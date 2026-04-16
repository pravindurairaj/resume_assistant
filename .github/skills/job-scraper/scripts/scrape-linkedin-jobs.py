#!/usr/bin/env python3
"""
scrape-linkedin-jobs.py - Scrape public LinkedIn job listings and rate fit against career profile.

Reads the user's career profile (.github/instructions/career-profile-{user}.instructions.md)
to auto-fill: keywords (Target Roles), location, and fit rating.

Usage:
    # Auto-read keywords & location from career profile (searches ALL target roles)
    python scrape-linkedin-jobs.py --user Pravin

    # Override keywords, location auto-filled from profile
    python scrape-linkedin-jobs.py -k "Data Analyst" --user Pravin

    # Multiple keywords in one run
    python scrape-linkedin-jobs.py -k "DevOps" -k "SRE" -k "Platform Engineer" --user Pravin

    # Full manual control
    python scrape-linkedin-jobs.py -k "DevOps Engineer" -l "London, UK" -d week --user Pravin

Requirements:
    pip install -r .github/requirements.txt
"""

import argparse
import re
import sys
import time
import urllib3
from datetime import datetime
from pathlib import Path

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ── Dependency check ──────────────────────────────────────────────────────────

def check_dependencies():
    missing = []
    try:
        import requests  # noqa: F401
    except ImportError:
        missing.append("requests")
    try:
        from bs4 import BeautifulSoup  # noqa: F401
    except ImportError:
        missing.append("beautifulsoup4")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print(f"ERROR: Missing dependencies: {', '.join(missing)}")
        print(f"Run:  pip install {' '.join(missing)}")
        print("Or:   pip install -r .github/requirements.txt")
        sys.exit(1)


# ── LinkedIn date-posted filter values ───────────────────────────────────────

TIME_FILTERS = {
    "day":   "r86400",
    "24h":   "r86400",
    "week":  "r604800",
    "month": "r2592000",
    "any":   "",
}


# ── Repo root detection ───────────────────────────────────────────────────────

def find_repo_root():
    """Walk up from script location to find repo root (contains .github/)."""
    path = Path(__file__).resolve().parent
    while path != path.parent:
        if (path / ".github").is_dir():
            return path
        path = path.parent
    return None


# ── Career profile loader (DRY: single function for all profile fields) ──────

def load_career_profile(user_name):
    """
    Load full career profile from .github/instructions/career-profile-{user}.instructions.md.

    Returns dict with keys: full_name, location, target_roles, target_industries.
    Returns empty dict if no profile found.
    """
    repo_root = find_repo_root()
    if not repo_root:
        return {}

    instructions_dir = repo_root / ".github" / "instructions"
    if not instructions_dir.exists():
        return {}

    for f in instructions_dir.glob("career-profile-*.instructions.md"):
        if user_name.lower() in f.stem.lower():
            content = f.read_text(encoding="utf-8")
            profile = {
                "full_name": "",
                "location": "",
                "target_roles": [],
                "target_industries": [],
            }

            current_section = ""
            for line in content.splitlines():
                stripped = line.strip()

                # Track section headers
                if stripped.startswith("## "):
                    current_section = stripped[3:].strip().lower()
                    continue

                # Extract from contact information table
                if current_section == "contact information" and "|" in stripped:
                    parts = [p.strip() for p in stripped.split("|")]
                    for i, part in enumerate(parts):
                        if "Location" in part and i + 1 < len(parts):
                            loc = parts[i + 1].strip().strip("*")
                            if loc and not loc.startswith("{"):
                                profile["location"] = loc
                        if "Full Name" in part and i + 1 < len(parts):
                            name = parts[i + 1].strip().strip("*")
                            if name and not name.startswith("{"):
                                profile["full_name"] = name

                # Extract list items from Target Roles
                if current_section == "target roles" and stripped.startswith("- "):
                    role = stripped[2:].strip()
                    if role and not role.startswith("{"):
                        profile["target_roles"].append(role)

                # Extract list items from Target Industries
                if current_section == "target industries" and stripped.startswith("- "):
                    ind = stripped[2:].strip()
                    if ind and not ind.startswith("{"):
                        profile["target_industries"].append(ind)

            return profile

    return {}


# ── LinkedIn scraper ──────────────────────────────────────────────────────────

def scrape_linkedin_jobs(keywords, location, date_posted="day", max_results=100):
    """
    Scrape public LinkedIn guest job search results.
    No login or API key required.
    """
    import requests
    from bs4 import BeautifulSoup

    time_filter = TIME_FILTERS.get(date_posted, "r86400")
    jobs = []

    url = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    for start in range(0, max_results, 25):
        params = {
            "keywords": keywords,
            "location": location,
            "start": start,
            "position": 1,
            "pageNum": 0,
        }
        if time_filter:
            params["f_TPR"] = time_filter

        try:
            resp = requests.get(url, params=params, headers=headers, timeout=15, verify=False)

            if resp.status_code == 429:
                print(f"  Rate limited (429). Waiting 30s before retrying...")
                time.sleep(30)
                resp = requests.get(url, params=params, headers=headers, timeout=15, verify=False)

            if resp.status_code != 200:
                print(f"  Page {start // 25 + 1}: HTTP {resp.status_code} — stopping.")
                break

            soup = BeautifulSoup(resp.text, "html.parser")
            cards = soup.find_all("div", class_="base-card")

            if not cards:
                print(f"  Page {start // 25 + 1}: no more results.")
                break

            for card in cards:
                title_el   = card.find("h3", class_="base-search-card__title")
                company_el = card.find("h4", class_="base-search-card__subtitle")
                link_el    = card.find("a", class_="base-card__full-link")
                loc_el     = card.find("span", class_="job-search-card__location")

                title    = title_el.get_text(strip=True)   if title_el   else ""
                company  = company_el.get_text(strip=True) if company_el else ""
                link     = link_el.get("href", "").split("?")[0] if link_el else ""
                job_loc  = loc_el.get_text(strip=True)     if loc_el     else ""

                if title and company:
                    jobs.append({
                        "title":    title,
                        "company":  company,
                        "link":     link,
                        "location": job_loc,
                    })

            print(f"  Page {start // 25 + 1}: {len(cards)} listings fetched  (total so far: {len(jobs)})")
            time.sleep(2)  # respectful rate limiting

        except requests.exceptions.Timeout:
            print(f"  Page {start // 25 + 1}: timeout — skipping page.")
            time.sleep(5)
            continue
        except Exception as e:
            print(f"  Page {start // 25 + 1}: unexpected error — {e}")
            break

    return jobs


# ── JD scraping (guest API) ─────────────────────────────────────────────────

def scrape_job_description(job_link):
    """
    Fetch full JD text from LinkedIn guest API for a single job posting.
    Extracts job_id from the link, fetches the public jobPosting endpoint,
    and parses the description HTML.

    Returns (jd_text: str, success: bool).
    Caller is responsible for rate-limiting (5s delay between calls).
    """
    import requests
    from bs4 import BeautifulSoup

    match = re.search(r'(\d{8,})', job_link or "")
    if not match:
        return "", False

    job_id = match.group(1)
    url = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{job_id}"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    try:
        resp = requests.get(url, headers=headers, timeout=15, verify=False)

        if resp.status_code == 429:
            print("    Rate limited (429). Waiting 30s...")
            time.sleep(30)
            resp = requests.get(url, headers=headers, timeout=15, verify=False)

        if resp.status_code != 200:
            return "", False

        soup = BeautifulSoup(resp.text, "html.parser")

        for selector in [
            "div.show-more-less-html__markup",
            "div.description__text",
            "section.show-more-less-html",
            "div.description",
        ]:
            el = soup.select_one(selector)
            if el:
                text = el.get_text(separator="\n", strip=True)
                if len(text) > 50:
                    return text, True

        return "", False

    except Exception:
        return "", False


# ── User skills loader ────────────────────────────────────────────────────────

def load_user_skills(user_name):
    """
    Read the user's master resume and extract all skills as a lowercase set.

    Parses:
    - '**Category:** Skill1, Skill2, ...' lines in the Skills section
    - '*Tech Stack: ...*' lines in Work Experience

    Returns set of lowercase skill strings.
    """
    repo_root = find_repo_root()
    if not repo_root:
        return set()

    resume_path = repo_root / ".github" / "Users" / user_name / f"{user_name}_Resume.md"
    if not resume_path.exists():
        return set()

    content = resume_path.read_text(encoding="utf-8")
    skills = set()

    for line in content.splitlines():
        stripped = line.strip()

        # Match: **Category:** Skill1, Skill2, Skill3
        m = re.match(r'\*\*[^*]+:\*\*\s*(.+)', stripped)
        if m:
            for s in re.split(r'[,;|·•]', m.group(1)):
                s = s.strip().strip('*').strip()
                if s and len(s) > 1:
                    skills.add(s.lower())

        # Match: *Tech Stack: Tool1, Tool2, ...*
        m = re.match(r'\*Tech Stack:\s*(.+?)\*?$', stripped)
        if m:
            for s in re.split(r'[,;|·•]', m.group(1)):
                s = s.strip().strip('*').strip()
                if s and len(s) > 1:
                    skills.add(s.lower())

    return skills


# ── Fit rating ────────────────────────────────────────────────────────────────

def calculate_fit(job_title, target_roles, jd_text="", user_skills=None):
    """
    Two-phase job fit scoring (0-100).

    Phase 1 (0-40 pts): Title match against target roles.
      - Exact phrase match = 40
      - Keyword coverage = proportional up to 40

    Phase 2 (0-60 pts): Tech/skill keyword extraction from JD text vs user skills.
      - Match ratio * 60

    Without JD text, max score is capped at 40 (title-only is unreliable).

    Returns (score: int, best_matching_role: str).
    """
    if not target_roles:
        return 0, ""

    STOP = {"a", "an", "the", "and", "or", "of", "in", "at", "to", "for",
            "is", "on", "with", "as", "be", "by"}

    job_words = set(re.findall(r"\w+", job_title.lower())) - STOP

    # ── Phase 1: Title match (0-40) ───────────────────────────────────────────
    best_title_score = 0
    best_role = ""

    for role in target_roles:
        role_words = set(re.findall(r"\w+", role.lower())) - STOP
        if not role_words or not job_words:
            continue

        if role.lower() in job_title.lower():
            best_title_score = 40
            best_role = role
            break

        matched  = len(role_words & job_words)
        coverage = matched / len(role_words)
        union    = job_words | role_words
        jaccard  = len(job_words & role_words) / len(union) if union else 0

        score = round(max(coverage, jaccard) * 40)
        if score > best_title_score:
            best_title_score = score
            best_role = role

    # ── Phase 2: JD skill match (0-60) ────────────────────────────────────────
    jd_skill_score = 0

    if jd_text and user_skills:
        jd_lower = jd_text.lower()

        SINGLE_TERMS = {
            "python", "java", "javascript", "typescript", "golang", "go", "rust",
            "ruby", "scala", "kotlin", "swift", "perl", "php", "r", "matlab",
            "sql", "nosql", "mysql", "postgresql", "postgres", "mongodb", "redis",
            "cassandra", "elasticsearch", "dynamodb", "oracle", "sqlite",
            "docker", "kubernetes", "k8s", "helm", "rancher", "openshift",
            "terraform", "ansible", "puppet", "chef", "vagrant",
            "jenkins", "gitlab", "github", "bitbucket", "bamboo", "circleci",
            "nexus", "artifactory", "sonarqube", "rundeck",
            "aws", "azure", "gcp", "cloudflare", "heroku", "digitalocean",
            "linux", "unix", "windows", "bash", "powershell", "zsh",
            "splunk", "datadog", "grafana", "prometheus", "nagios", "elk",
            "kafka", "rabbitmq", "celery", "airflow",
            "pandas", "numpy", "scipy", "matplotlib", "seaborn", "plotly",
            "scikit-learn", "tensorflow", "pytorch", "keras",
            "pyspark", "spark", "hadoop", "hive", "flink",
            "tableau", "looker", "qlik",
            "jira", "confluence", "servicenow", "pagerduty",
            "git", "svn", "mercurial",
            "nginx", "apache", "tomcat", "iis",
            "react", "angular", "vue", "node", "express", "flask", "django",
            "spring", "dotnet", "csharp", "vbscript",
            "hl7", "fhir", "dicom", "blockchain", "solidity",
            "excel", "putty", "winscp", "stackstorm",
        }

        MULTI_TERMS = {
            "machine learning", "deep learning", "natural language processing",
            "shell scripting", "ci/cd", "ci cd", "power bi", "power automate",
            "amazon web services", "google cloud", "azure devops",
            "site reliability", "data engineering", "data analysis",
            "agile scrum", "azure machine learning", "splunk enterprise",
            "splunk cloud", "r programming",
        }

        jd_keywords = set()
        for term in SINGLE_TERMS:
            if re.search(r'\b' + re.escape(term) + r'\b', jd_lower):
                jd_keywords.add(term)
        for term in MULTI_TERMS:
            if term in jd_lower:
                jd_keywords.add(term)

        if jd_keywords:
            matched = jd_keywords & user_skills
            for jd_kw in jd_keywords - matched:
                for us in user_skills:
                    if jd_kw in us or us in jd_kw:
                        matched.add(jd_kw)
                        break
            ratio = len(matched) / len(jd_keywords)
            jd_skill_score = round(ratio * 60)

    # ── Combine ───────────────────────────────────────────────────────────────
    if not jd_text:
        total = best_title_score
    else:
        total = best_title_score + jd_skill_score

    return total, best_role


# ── Excel generation ──────────────────────────────────────────────────────────

def save_to_excel(jobs, output_path, target_roles=None, search_meta=None):
    """Write scraped jobs to a professionally styled .xlsx file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Jobs"

    # ── Header metadata row ────────────────────────────────────────────────────
    if search_meta:
        ws.append([f"LinkedIn Job Search — {search_meta}"])
        ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
        ws.append([f"Generated: {datetime.now().strftime('%d %b %Y %H:%M:%S')}"])
        ws["A2"].font = Font(italic=True, size=10, color="555555")
        ws.append([])  # spacer
        header_row_num = 4
    else:
        header_row_num = 1

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["#", "Company", "Job Title", "Location", "Job Link"]
    if target_roles:
        headers += ["Fit %", "Best Matching Role"]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),   right=Side(style="thin"),
        top=Side(style="thin"),    bottom=Side(style="thin"),
    )

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row_num, column=col)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border
    ws.row_dimensions[header_row_num].height = 22

    # ── Sort by pre-computed fit (scoring moved to main) ────────────────────
    if target_roles:
        jobs.sort(key=lambda j: j.get("_fit", 0), reverse=True)

    # Color fills for fit rating
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amber = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Alternating row fills for readability
    fill_row_a = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    fill_row_b = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ── Data rows ──────────────────────────────────────────────────────────────
    for idx, job in enumerate(jobs, 1):
        row_num = header_row_num + idx
        row_fill = fill_row_a if idx % 2 == 0 else fill_row_b

        def cell(col, value, _rn=row_num, _rf=row_fill):
            c = ws.cell(row=_rn, column=col, value=value)
            c.border = thin_border
            c.fill   = _rf
            return c

        cell(1, idx).alignment = Alignment(horizontal="center")
        cell(2, job["company"])
        cell(3, job["title"])
        cell(4, job.get("location", ""))

        # Clickable hyperlink
        lc = cell(5, job["link"] or "")
        if job["link"]:
            lc.hyperlink = job["link"]
            lc.value     = "Open Job ↗"
        lc.font      = Font(color="1F4E79", underline="single")
        lc.alignment = Alignment(horizontal="center")

        if target_roles:
            fit = job.get("_fit", 0)
            fc  = cell(6, f"{fit}%")
            fc.alignment = Alignment(horizontal="center")
            fc.font      = Font(bold=True)
            fc.fill      = fill_green if fit >= 75 else (fill_amber if fit >= 50 else fill_red)

            cell(7, job.get("_role", ""))

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = {"A": 5, "B": 30, "C": 40, "D": 22, "E": 14, "F": 10, "G": 32}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze panes at first data row
    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=1)

    # ── Summary sheet ──────────────────────────────────────────────────────────
    if target_roles and jobs:
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15

        summary_headers = ["Metric", "Count"]
        for ci, h in enumerate(summary_headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

        high  = sum(1 for j in jobs if j.get("_fit", 0) >= 75)
        med   = sum(1 for j in jobs if 50 <= j.get("_fit", 0) < 75)
        low   = sum(1 for j in jobs if j.get("_fit", 0) < 50)
        total = len(jobs)

        summary_data = [
            ("Total Jobs Found",         total),
            ("High Fit (≥75%)",          high),
            ("Medium Fit (50–74%)",      med),
            ("Low Fit (<50%)",           low),
        ]
        for ri, (label, val) in enumerate(summary_data, 2):
            ws2.cell(row=ri, column=1, value=label).border = thin_border
            vc = ws2.cell(row=ri, column=2, value=val)
            vc.border    = thin_border
            vc.alignment = Alignment(horizontal="center")

    # ── Save ───────────────────────────────────────────────────────────────────
    wb.properties.creator  = "Job Scraper"
    wb.properties.keywords = "LinkedIn Jobs Resume Fit"
    wb.save(str(output_path))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Scrape LinkedIn public job listings and rate fit against career profile."
    )
    parser.add_argument("--keywords",    "-k", action="append", default=None,
                        help="Search keywords (repeatable). If omitted, uses Target Roles from career profile.")
    parser.add_argument("--location",    "-l", default=None,
                        help="Location filter. If omitted, uses Location from career profile.")
    parser.add_argument("--date-posted", "-d", default="day",
                        choices=["day", "24h", "week", "month", "any"],
                        help="Date posted filter (default: day/24h)")
    parser.add_argument("--user",        "-u", default="Pravin",
                        help="User name — reads career profile for location, keywords, fit (default: Pravin)")
    parser.add_argument("--max-results", "-m", type=int, default=100,
                        help="Max listings per keyword (default: 100)")
    parser.add_argument("--output-dir",  "-o", default=None,
                        help="Output directory (default: {UserName}/JobSearch)")

    args = parser.parse_args()
    check_dependencies()

    # ── Load career profile ────────────────────────────────────────────────────
    profile = load_career_profile(args.user)
    target_roles     = profile.get("target_roles", []) if profile else []
    profile_location = profile.get("location", "")    if profile else ""

    if profile:
        print(f"User        : {args.user}")
        if profile.get("full_name"):
            print(f"Full name   : {profile['full_name']}")
        if profile_location:
            print(f"Location    : {profile_location}")
        if target_roles:
            print(f"Target roles: {', '.join(target_roles)}")
    else:
        print(f"No career profile found for '{args.user}'. Using manual args only.")

    # ── Resolve keywords: CLI > career profile target roles ────────────────────
    if args.keywords:
        keyword_list = args.keywords
    elif target_roles:
        keyword_list = target_roles
        print(f"\nNo --keywords provided. Using all {len(keyword_list)} target roles as keywords.")
    else:
        print("\nERROR: No --keywords and no target roles in career profile. Provide at least one -k.")
        sys.exit(1)

    # ── Resolve location: CLI > career profile ─────────────────────────────────
    location    = args.location if args.location is not None else profile_location
    loc_display = location or "Any location"

    # ── Resolve output dir: CLI > {UserName}/JobSearch ─────────────────────────
    out_dir = Path(args.output_dir) if args.output_dir else Path(args.user) / "JobSearch"
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Scrape all keywords ────────────────────────────────────────────────────
    all_jobs: list   = []
    seen_links: set  = set()  # deduplicate across keyword runs

    for kw in keyword_list:
        print(f"\n{'='*60}")
        print(f"Searching : '{kw}' in '{loc_display}' (posted: {args.date_posted})")
        print(f"{'='*60}\n")

        jobs = scrape_linkedin_jobs(kw, location, args.date_posted, args.max_results)

        new_jobs = []
        for j in jobs:
            key = j["link"] or f"{j['company']}|{j['title']}"
            if key not in seen_links:
                seen_links.add(key)
                j["_search_keyword"] = kw
                new_jobs.append(j)

        all_jobs.extend(new_jobs)
        dupes = len(jobs) - len(new_jobs)
        print(f"  → {len(new_jobs)} new unique jobs  (duplicates skipped: {dupes})")

    if not all_jobs:
        print("\nNo jobs found. Try: broader keywords, different date filter, or check connectivity.")
        sys.exit(0)

    print(f"\n{'='*60}")
    print(f"Total unique jobs: {len(all_jobs)}")
    print(f"{'='*60}")

    # ── Load user skills for JD matching ──────────────────────────────────────
    user_skills = load_user_skills(args.user)
    if user_skills:
        print(f"\nLoaded {len(user_skills)} skills from {args.user}'s master resume.")
    else:
        print(f"\nWARNING: No skills loaded for '{args.user}' — JD skill matching disabled.")

    # ── Scrape JDs from LinkedIn guest API ────────────────────────────────────
    print(f"\nFetching JD text for {len(all_jobs)} job(s) via LinkedIn guest API...")
    for i, j in enumerate(all_jobs, 1):
        label = f"{j['company'][:25]} — {j['title'][:30]}"
        print(f"  [{i}/{len(all_jobs)}] {label}...", end=" ", flush=True)
        jd_text, ok = scrape_job_description(j["link"])
        j["_jd_text"] = jd_text
        print("\u2713" if ok else "\u2717")
        if i < len(all_jobs):
            time.sleep(5)

    # ── Two-phase fit scoring (title + JD skills) ─────────────────────────────
    if target_roles:
        for j in all_jobs:
            j["_fit"], j["_role"] = calculate_fit(
                j["title"], target_roles, j.get("_jd_text", ""), user_skills
            )
        all_jobs.sort(key=lambda j: j["_fit"], reverse=True)

    # ── Output filename ────────────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if len(keyword_list) == 1:
        safe_kw  = re.sub(r"[^\w\-]", "_", keyword_list[0]).strip("_")
        filename = f"LinkedIn_Jobs_{safe_kw}_{ts}.xlsx"
    else:
        filename = f"LinkedIn_Jobs_AllRoles_{ts}.xlsx"

    output_path = out_dir / filename

    kw_display  = ", ".join(keyword_list)
    search_meta = f"{kw_display} · {loc_display} · {args.date_posted}"
    save_to_excel(
        all_jobs,
        output_path,
        target_roles=target_roles if target_roles else None,
        search_meta=search_meta,
    )

    # ── Summary ────────────────────────────────────────────────────────────────
    print(f"\nSaved : {output_path}")
    if target_roles:
        high = sum(1 for j in all_jobs if j.get("_fit", 0) >= 75)
        med  = sum(1 for j in all_jobs if 50 <= j.get("_fit", 0) < 75)
        low  = len(all_jobs) - high - med
        print(f"Fit   : 🟢 High ≥75%: {high}  🟡 Medium 50-74%: {med}  🔴 Low <50%: {low}")
    print(f"Open  : {output_path}")


if __name__ == "__main__":
    main()
